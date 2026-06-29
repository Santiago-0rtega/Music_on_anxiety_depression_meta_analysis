from __future__ import annotations

import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "data" / "MUSIC-CRIME-Q_RoB_assessment.xlsx"
CODEBOOK_PATH = ROOT / "CRIMEQ" / "notebooklm_workflow" / "shared_rules" / "CRIME-Q_CODEBOOK_FOR_NOTEBOOKLM.md"
ASSESSMENT_SHEET = "Sheet1"
CODEBOOK_SHEET = "CRIME-Q Codebook"

OPTION_COLUMNS = ["Yes", "Partly", "No", "Unclear", "NA"]
CODEBOOK_HEADERS = [
    "Item",
    "Item name",
    "Question",
    "Response options",
    *OPTION_COLUMNS,
    "Decision rule",
]


def compact_text(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip())


def extract_block(text: str, start: str, stop: str) -> str:
    pattern = re.compile(rf"{re.escape(start)}\s*(.*?)(?=\n\n{re.escape(stop)})", re.S)
    match = pattern.search(text)
    return compact_text(match.group(1)) if match else ""


def parse_bullets(text: str) -> dict[str, str]:
    bullets: dict[str, str] = {}
    current_label: str | None = None

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        bullet_match = re.match(r"^-\s+([^:]+):\s*(.*)$", line)
        if bullet_match:
            label = bullet_match.group(1).strip()
            value = bullet_match.group(2).strip()
            if label in OPTION_COLUMNS or label == "Decision rule":
                bullets[label] = value
                current_label = label
            else:
                current_label = None
            continue

        if current_label and line.startswith("  ") and line.strip():
            bullets[current_label] = compact_text(f"{bullets[current_label]} {line.strip()}")

    return bullets


def parse_codebook() -> list[dict[str, str]]:
    source = CODEBOOK_PATH.read_text(encoding="utf-8")
    headings = list(re.finditer(r"^##\s+(.+?)\s+-\s+(.+?)\s*$", source, re.M))
    entries: list[dict[str, str]] = []

    for idx, match in enumerate(headings):
        item = match.group(1).strip()
        item_name = match.group(2).strip()
        start = match.end()
        end = headings[idx + 1].start() if idx + 1 < len(headings) else len(source)
        section = source[start:end].strip()

        response_match = re.search(r"Response options:\s*(.*?)(?=\n\n-)", section, re.S)
        response_options = compact_text(response_match.group(1)) if response_match else ""
        bullets = parse_bullets(section)

        entry = {
            "Item": item,
            "Item name": item_name,
            "Question": extract_block(section, "Question:", "Response options:"),
            "Response options": response_options,
            "Decision rule": bullets.get("Decision rule", ""),
        }
        for option in OPTION_COLUMNS:
            entry[option] = bullets.get(option, "")
        entries.append(entry)

    return entries


def score_items_from_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    items: list[str] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if not value:
            continue
        header = str(value).strip()
        if header in {"Study", "Study_Title"}:
            continue
        if header.endswith("_JUSTIFICATION") or header.endswith("_VERBATIM"):
            continue
        items.append(header)
    return items


def header_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1)}


def make_header_comment(entry: dict[str, str]) -> str:
    lines = [
        f"{entry['Item']} - {entry['Item name']}",
        "",
        f"Question: {entry['Question']}",
        f"Response options: {entry['Response options']}",
        "",
    ]
    for option in OPTION_COLUMNS:
        if entry.get(option):
            lines.append(f"{option}: {entry[option]}")
    if entry.get("Decision rule"):
        lines.extend(["", f"Decision rule: {entry['Decision rule']}"])
    lines.extend(["", f"Source: {CODEBOOK_PATH.name}"])
    return "\n".join(lines)


def add_header_comments(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    entries_by_item: dict[str, dict[str, str]],
) -> None:
    headers = header_columns(ws)
    for item, entry in entries_by_item.items():
        if item not in headers:
            continue
        cell = ws.cell(1, headers[item])
        cell.comment = Comment(make_header_comment(entry), "Codex")
        cell.comment.width = 420
        cell.comment.height = 300


def style_codebook_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, row_count: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="FF1F4E78")
    header_font = Font(bold=True, color="FFFFFFFF")
    thin_gray = Side(style="thin", color="FFD9E2EC")
    border = Border(bottom=thin_gray)

    for cell in ws[1]:
        cell.fill = copy(header_fill)
        cell.font = copy(header_font)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(border)

    for row in ws.iter_rows(min_row=2, max_row=row_count, min_col=1, max_col=len(CODEBOOK_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = copy(border)

    widths = {
        "A": 12,
        "B": 34,
        "C": 58,
        "D": 22,
        "E": 52,
        "F": 52,
        "G": 52,
        "H": 52,
        "I": 36,
        "J": 58,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 30
    for row_num in range(2, row_count + 1):
        ws.row_dimensions[row_num].height = 120

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{row_count}"

    table = Table(displayName="CrimeQCodebook", ref=f"A1:J{row_count}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def rebuild_codebook_sheet(
    wb: openpyxl.Workbook,
    entries: list[dict[str, str]],
    item_order: list[str],
) -> None:
    if CODEBOOK_SHEET in wb.sheetnames:
        del wb[CODEBOOK_SHEET]

    ws = wb.create_sheet(CODEBOOK_SHEET)
    entries_by_item = {entry["Item"]: entry for entry in entries}
    ordered_entries = [entries_by_item[item] for item in item_order if item in entries_by_item]

    ws.append(CODEBOOK_HEADERS)
    for entry in ordered_entries:
        ws.append([entry.get(header, "") for header in CODEBOOK_HEADERS])

    style_codebook_sheet(ws, len(ordered_entries) + 1)


def main() -> None:
    entries = parse_codebook()
    entries_by_item = {entry["Item"]: entry for entry in entries}

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    assessment_ws = wb[ASSESSMENT_SHEET]
    item_order = score_items_from_header(assessment_ws)

    missing_from_codebook = [item for item in item_order if item not in entries_by_item]
    if missing_from_codebook:
        raise RuntimeError(f"Assessment headers missing from codebook: {missing_from_codebook}")

    add_header_comments(assessment_ws, entries_by_item)
    rebuild_codebook_sheet(wb, entries, item_order)
    wb.save(WORKBOOK_PATH)

    print(f"Workbook updated: {WORKBOOK_PATH}")
    print(f"Header comments refreshed: {len(item_order)}")
    print(f"Codebook rows written: {len(item_order)}")
    print(f"Codebook sheet: {CODEBOOK_SHEET}")


if __name__ == "__main__":
    main()
