from __future__ import annotations

import json
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[2]
CURRENT_WORKBOOK = ROOT / "data" / "MUSIC-CRIME-Q_RoB_assessment.xlsx"
GOLD_WORKBOOK = ROOT / "data" / "MUSIC-CRIME-Q_GOLD-STANDARD_assessment.xlsx"
DISAGREEMENTS_JSON = ROOT / "CRIMEQ" / "independent_screening" / "current_disagreements.json"

COMMENT_PREFIX = "CRIME-Q comparison note"

SCORE_FILL = {
    "Yes": PatternFill(fill_type="solid", fgColor="FFC6EFCE"),
    "No": PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
    "Partly": PatternFill(fill_type="solid", fgColor="FFFFEB9C"),
    "Unclear": PatternFill(fill_type="solid", fgColor="FFF4B183"),
}

CANONICAL_SCORE = {
    "yes": "Yes",
    "no": "No",
    "partly": "Partly",
    "partial": "Partly",
    "unclear": "Unclear",
}


AGREED_OVERRIDES = {
    ("Li2010", "10Z"): {
        "score": "Yes",
        "justification": (
            "Public/academic funding was reported, with no commercial funder or funder-control "
            "signal found."
        ),
        "verbatim": (
            '[p.77, Acknowledgments] "This study was supported by Natural Science Foundation '
            'of China (no. 30725020, 30700258)..."'
        ),
    },
    ("Pangemanan2024", "5Y"): {
        "score": "Partly",
        "justification": (
            "Music/no-music exposure used separate rooms and rats were housed two per cage, "
            "creating room and cage/pseudoreplication concerns, but not a clear one-cage, "
            "same-litter, or same-cage fatal confound."
        ),
        "verbatim": (
            '[p.349, Methods] "The experimental animals were collectively housed, with two '
            'rats accommodated in each cage..." and "A separate room was designated for '
            'groups without music exposure..."'
        ),
    },
    ("Saghari2021", "9X"): {
        "score": "Partly",
        "justification": (
            "The authors briefly identify that the mechanism of music's effect was not studied "
            "and call for further mechanistic work; this is a study-specific limitation, "
            "though brief."
        ),
        "verbatim": (
            '[p.7, Conclusions] "the mechanism by which music alleviated the impairments '
            'induced by SPS in rats is not studied. Further studies... are needed"'
        ),
    },
    ("Sampaio2017", "5X"): {
        "score": "Yes",
        "justification": (
            "The report gives the stimulus, intensity, exposure schedule, duration, source/file "
            "preparation, playback device, device placement, and sound-balancing checks."
        ),
        "verbatim": (
            '[p.177, Music Therapy] "Mozart\'s Sonata for Two Pianos... 65 decibels... '
            'sound-emitting device... approximately 50 cm from the rats\' cages... sound '
            'intensity was measured on the sides of each cage."'
        ),
    },
    ("Sampaio2017", "9X"): {
        "score": "Yes",
        "justification": (
            "The authors explicitly discuss study-specific caveats, including unmeasured "
            "hormones and possible adaptation/test-repetition effects."
        ),
        "verbatim": (
            '[p.184-186, Discussion/Conclusions] "hormone levels... were not measured in our '
            'study" and "it cannot be discarded that these effects might have been due to '
            'the animals\' adaptation to the tests..."'
        ),
    },
}


def canonical_score(value: object) -> str:
    if value is None:
        return ""
    key = str(value).strip().lower()
    return CANONICAL_SCORE.get(key, str(value).strip())


def compact_study_id(value: object) -> str:
    text = str(value).strip()
    parts = text.split("_")
    if len(parts) >= 2 and parts[1].isdigit():
        return f"{parts[0]}{parts[1]}"
    return text


def headers_by_name(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1)}


def study_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        study = ws.cell(row, 1).value
        if study:
            rows[str(study).strip()] = row
    return rows


def read_gold() -> dict[tuple[str, str], dict[str, str]]:
    wb = openpyxl.load_workbook(GOLD_WORKBOOK, data_only=False)
    ws = wb["Full_Assessment"]
    out: dict[tuple[str, str], dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        study = compact_study_id(ws.cell(row, 1).value)
        item = str(ws.cell(row, 2).value).strip()
        out[(study, item)] = {
            "score": canonical_score(ws.cell(row, 4).value),
            "justification": ws.cell(row, 5).value or "",
            "verbatim": ws.cell(row, 6).value or "",
        }
    return out


def score_columns(header: dict[str, int]) -> dict[str, int]:
    columns: dict[str, int] = {}
    for name, col in header.items():
        if name in {"Study", "Study_Title"}:
            continue
        if name.endswith("_JUSTIFICATION") or name.endswith("_VERBATIM"):
            continue
        columns[name] = col
    return columns


def set_wrap(cell: openpyxl.cell.cell.Cell) -> None:
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "top"
    cell.alignment = alignment


def main() -> None:
    with DISAGREEMENTS_JSON.open(encoding="utf-8") as fh:
        disagreements = json.load(fh)

    gold = read_gold()
    wb = openpyxl.load_workbook(CURRENT_WORKBOOK)
    ws = wb["Sheet1"]
    header = headers_by_name(ws)
    rows = study_rows(ws)
    score_cols = score_columns(header)

    missing = []
    applied = []

    for item, col in score_cols.items():
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            score = canonical_score(cell.value)
            if score in SCORE_FILL:
                cell.value = score
                cell.fill = copy(SCORE_FILL[score])
            if cell.comment and cell.comment.text.startswith(COMMENT_PREFIX):
                cell.comment = None

    for disagreement in disagreements:
        study = str(disagreement["study"]).strip()
        item = str(disagreement["item"]).strip()
        key = (study, item)

        if study not in rows or item not in score_cols:
            missing.append({"study": study, "item": item, "reason": "missing current workbook location"})
            continue

        final = AGREED_OVERRIDES.get(key, gold.get(key))
        if not final:
            missing.append({"study": study, "item": item, "reason": "missing gold/override record"})
            continue

        row = rows[study]
        score_col = score_cols[item]
        just_col = header.get(f"{item}_JUSTIFICATION")
        verb_col = header.get(f"{item}_VERBATIM")
        if just_col is None or verb_col is None:
            missing.append({"study": study, "item": item, "reason": "missing adjacent detail columns"})
            continue

        score_cell = ws.cell(row, score_col)
        old_score = canonical_score(score_cell.value)
        new_score = canonical_score(final["score"])
        score_cell.value = new_score
        score_cell.fill = copy(SCORE_FILL[new_score])
        if score_cell.comment and score_cell.comment.text.startswith(COMMENT_PREFIX):
            score_cell.comment = None

        just_cell = ws.cell(row, just_col)
        verb_cell = ws.cell(row, verb_col)
        just_cell.value = final["justification"]
        verb_cell.value = final["verbatim"]
        set_wrap(just_cell)
        set_wrap(verb_cell)

        applied.append(
            {
                "study": study,
                "item": item,
                "old_score": old_score,
                "new_score": new_score,
                "source": "agreed_override" if key in AGREED_OVERRIDES else "gold_standard",
                "score_cell": score_cell.coordinate,
                "justification_cell": just_cell.coordinate,
                "verbatim_cell": verb_cell.coordinate,
            }
        )

    header_comments = 0
    data_score_comments = 0
    fill_counts = {key: 0 for key in SCORE_FILL}
    invalid_scores = []
    nonblank_details = 0

    for item, col in score_cols.items():
        if ws.cell(1, col).comment:
            header_comments += 1
        for row in range(2, ws.max_row + 1):
            score = canonical_score(ws.cell(row, col).value)
            if score in SCORE_FILL:
                fill_counts[score] += 1
            else:
                invalid_scores.append({"cell": ws.cell(row, col).coordinate, "value": ws.cell(row, col).value})
            if ws.cell(row, col).comment:
                data_score_comments += 1

    for entry in applied:
        just_value = ws[entry["justification_cell"]].value
        verb_value = ws[entry["verbatim_cell"]].value
        if just_value and verb_value:
            nonblank_details += 1

    wb.save(CURRENT_WORKBOOK)

    print(f"Workbook updated: {CURRENT_WORKBOOK}")
    print(f"Applied reviewed cells: {len(applied)}")
    print(f"Applied from gold standard: {sum(1 for x in applied if x['source'] == 'gold_standard')}")
    print(f"Applied agreed overrides: {sum(1 for x in applied if x['source'] == 'agreed_override')}")
    print(f"Header comments retained: {header_comments}")
    print(f"Data score comments remaining: {data_score_comments}")
    print(f"Revised cells with justification and verbatim: {nonblank_details}/{len(applied)}")
    print(f"Fill counts: {fill_counts}")
    print(f"Invalid/uncolored score cells: {len(invalid_scores)}")
    if missing:
        print("Missing records:")
        print(json.dumps(missing, indent=2))
        raise SystemExit(1)


if __name__ == "__main__":
    main()
