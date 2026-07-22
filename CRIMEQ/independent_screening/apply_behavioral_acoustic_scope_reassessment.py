from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "data" / "MUSIC-CRIME-Q_RoB_assessment.xlsx"
ASSESSMENT_SHEET = "Sheet1"

SCORE_FILLS = {
    "Yes": PatternFill(fill_type="solid", fgColor="FFC6EFCE"),
    "No": PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
    "Partly": PatternFill(fill_type="solid", fgColor="FFFFEB9C"),
    "Unclear": PatternFill(fill_type="solid", fgColor="FFF4B183"),
    "NA": PatternFill(fill_type="solid", fgColor="FF9DC3E6"),
}

REVISIONS = {
    ("Freitas2020", "5X"): {
        "score": "Partly",
        "justification": (
            "Music type/composers, intensity, schedule, duration, and control "
            "condition are reported, but playback hardware and cage/room "
            "delivery arrangement are not described."
        ),
        "verbatim": (
            '[p.13] "Diverse musical pieces of various classical music composers '
            'including Bach, Chopin... Mozart... reproduced around 65dB intensity '
            'levels... remained for 12 hours... three times a week during the '
            'dark cycle, three months on alternated days."'
        ),
    },
    ("Li2010", "7X"): {
        "score": "Partly",
        "justification": (
            "Behavioral rater blinding is reported, but only to genotype; "
            "blinding to acoustic/music exposure is not stated."
        ),
        "verbatim": (
            '[Methods] "All behavioral measurements were performed by raters '
            'blind to genotype."'
        ),
    },
    ("Cheng2024", "8Z(1)"): {
        "score": "Yes",
        "justification": (
            "Behavioral group size is stated and behavioral figures report "
            "N=11 per group; smaller molecular subsets are outside the "
            "behavioral attrition item under the revised scope."
        ),
        "verbatim": (
            '[Fig. 1] "N = 11 in each group" and [Fig. 3] "N = 4 - 6 in '
            'each group" (molecular figure only).'
        ),
    },
    ("Krishnamurthy2025", "9X"): {
        "score": "No",
        "justification": (
            "The limitations/future-work paragraph focuses on mechanistic "
            "biomarker and histology studies, not behavioral assay feasibility, "
            "acoustic delivery/control, or behavioral interpretation under the "
            "revised scope."
        ),
        "verbatim": (
            '[p.326] "Further studies include, estimation of cortisol levels, '
            'neurotransmitter levels, Brain-derived neurotrophic factor (BDNF), '
            'cytokine profiling and histological studies..."'
        ),
    },
    ("Saghari2021", "9X"): {
        "score": "No",
        "justification": (
            "The stated limitation is that the neural mechanism was not studied; "
            "no limitation is given for behavioral assays, acoustic exposure/"
            "control, or behavioral interpretation under the revised scope."
        ),
        "verbatim": (
            '[p.7] "the mechanism by which music alleviated the impairments '
            'induced by SPS in rats is not studied."'
        ),
    },
    ("Sampaio2017", "9X"): {
        "score": "Partly",
        "justification": (
            "One behavioral-assay limitation remains: possible adaptation from "
            "repeating the same tests at later developmental stages. The "
            "unmeasured hormone caveat is mechanistic and no longer drives "
            "this item."
        ),
        "verbatim": (
            '[p.186] "it cannot be discarded that these effects might have '
            "been due to the animals' adaptation to the tests because the same "
            'test was reapplied at each developmental stage."'
        ),
    },
}


def header_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1)}


def study_rows(ws: openpyxl.worksheet.worksheet.Worksheet, study_col: int) -> dict[str, int]:
    return {str(ws.cell(row, study_col).value).strip(): row for row in range(2, ws.max_row + 1)}


def apply_revisions() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb[ASSESSMENT_SHEET]
    headers = header_columns(ws)
    rows = study_rows(ws, headers["Study"])

    for (study, item), revision in REVISIONS.items():
        if study not in rows:
            raise RuntimeError(f"Study not found: {study}")
        for required in (item, f"{item}_JUSTIFICATION", f"{item}_VERBATIM"):
            if required not in headers:
                raise RuntimeError(f"Column not found: {required}")

        row = rows[study]
        score = revision["score"]
        score_cell = ws.cell(row, headers[item])
        score_cell.value = score
        score_cell.fill = copy(SCORE_FILLS[score])

        ws.cell(row, headers[f"{item}_JUSTIFICATION"]).value = revision["justification"]
        ws.cell(row, headers[f"{item}_VERBATIM"]).value = revision["verbatim"]

    wb.save(WORKBOOK_PATH)
    print(f"Updated workbook: {WORKBOOK_PATH}")
    print(f"Revised cells: {len(REVISIONS)}")
    for study, item in REVISIONS:
        print(f"- {study} {item} -> {REVISIONS[(study, item)]['score']}")


if __name__ == "__main__":
    apply_revisions()
