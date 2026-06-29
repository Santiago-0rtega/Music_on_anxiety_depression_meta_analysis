from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, PatternFill


ROOT = Path(__file__).resolve().parents[2]
CURRENT_WORKBOOK = ROOT / "data" / "MUSIC-CRIME-Q_RoB_assessment.xlsx"
GOLD_WORKBOOK = ROOT / "data" / "MUSIC-CRIME-Q_GOLD-STANDARD_assessment.xlsx"

SCORE_FILL = {
    "Yes": PatternFill(fill_type="solid", fgColor="FFC6EFCE"),
    "No": PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
    "Partly": PatternFill(fill_type="solid", fgColor="FFFFEB9C"),
    "Unclear": PatternFill(fill_type="solid", fgColor="FFF4B183"),
}

CANONICAL_SCORE = {
    "yes": "Yes",
    "no": "No",
    "partly": "Partly",
    "partial": "Partly",
    "unclear": "Unclear",
}

AGREED_OVERRIDES = {
    ("Li2010", "10Z"): {
        "score": "Yes",
        "justification": (
            "Public/academic funding was reported, with no commercial funder or funder-control "
            "signal found."
        ),
        "verbatim": (
            '[p.77, Acknowledgments] "This study was supported by Natural Science Foundation '
            'of China (no. 30725020, 30700258)..."'
        ),
    },
    ("Pangemanan2024", "5Y"): {
        "score": "Partly",
        "justification": (
            "Music/no-music exposure used separate rooms and rats were housed two per cage, "
            "creating room and cage/pseudoreplication concerns, but not a clear one-cage, "
            "same-litter, or same-cage fatal confound."
        ),
        "verbatim": (
            '[p.349, Methods] "The experimental animals were collectively housed, with two '
            'rats accommodated in each cage..." and "A separate room was designated for '
            'groups without music exposure..."'
        ),
    },
    ("Saghari2021", "9X"): {
        "score": "Partly",
        "justification": (
            "The authors briefly identify that the mechanism of music's effect was not studied "
            "and call for further mechanistic work; this is a study-specific limitation, "
            "though brief."
        ),
        "verbatim": (
            '[p.7, Conclusions] "the mechanism by which music alleviated the impairments '
            'induced by SPS in rats is not studied. Further studies... are needed"'
        ),
    },
    ("Sampaio2017", "5X"): {
        "score": "Yes",
        "justification": (
            "The report gives the stimulus, intensity, exposure schedule, duration, source/file "
            "preparation, playback device, device placement, and sound-balancing checks."
        ),
        "verbatim": (
            '[p.177, Music Therapy] "Mozart\'s Sonata for Two Pianos... 65 decibels... '
            'sound-emitting device... approximately 50 cm from the rats\' cages... sound '
            'intensity was measured on the sides of each cage."'
        ),
    },
    ("Sampaio2017", "9X"): {
        "score": "Yes",
        "justification": (
            "The authors explicitly discuss study-specific caveats, including unmeasured "
            "hormones and possible adaptation/test-repetition effects."
        ),
        "verbatim": (
            '[p.184-186, Discussion/Conclusions] "hormone levels... were not measured in our '
            'study" and "it cannot be discarded that these effects might have been due to '
            'the animals\' adaptation to the tests..."'
        ),
    },
}


def canonical_score(value: object) -> str:
    if value is None:
        return ""
    key = str(value).strip().lower()
    return CANONICAL_SCORE.get(key, str(value).strip())


def compact_study_id(value: object) -> str:
    text = str(value).strip()
    parts = text.split("_")
    if len(parts) >= 2 and parts[1].isdigit():
        return f"{parts[0]}{parts[1]}"
    return text


def read_gold_full() -> dict[tuple[str, str], dict[str, str]]:
    wb = openpyxl.load_workbook(GOLD_WORKBOOK, data_only=False)
    ws = wb["Full_Assessment"]
    records: dict[tuple[str, str], dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        key = (compact_study_id(ws.cell(row, 1).value), str(ws.cell(row, 2).value).strip())
        records[key] = {
            "score": canonical_score(ws.cell(row, 4).value),
            "justification": ws.cell(row, 5).value or "",
            "verbatim": ws.cell(row, 6).value or "",
        }
    records.update(AGREED_OVERRIDES)
    return records


def score_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    columns: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if not value:
            continue
        header = str(value).strip()
        if header in {"Study", "Study_Title"}:
            continue
        if header.endswith("_JUSTIFICATION") or header.endswith("_VERBATIM"):
            continue
        columns[header] = col
    return columns


def headers_by_name(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1)}


def restore() -> None:
    target = read_gold_full()
    wb = openpyxl.load_workbook(CURRENT_WORKBOOK)
    ws = wb["Sheet1"]
    headers = headers_by_name(ws)
    score_cols = score_columns(ws)

    changed_scores = 0
    detail_cells_written = 0
    missing: list[tuple[str, str]] = []

    for row in range(2, ws.max_row + 1):
        study = str(ws.cell(row, headers["Study"]).value).strip()
        for item, score_col in score_cols.items():
            key = (study, item)
            record = target.get(key)
            if not record:
                missing.append(key)
                continue

            score = canonical_score(record["score"])
            score_cell = ws.cell(row, score_col)
            if canonical_score(score_cell.value) != score:
                changed_scores += 1
            score_cell.value = score
            if score in SCORE_FILL:
                score_cell.fill = copy(SCORE_FILL[score])

            for suffix, field in [("_JUSTIFICATION", "justification"), ("_VERBATIM", "verbatim")]:
                detail_col = headers[f"{item}{suffix}"]
                detail_cell = ws.cell(row, detail_col)
                detail_cell.value = record[field]
                detail_cell.alignment = Alignment(vertical="top", wrap_text=True)
                detail_cells_written += 1

    if missing:
        raise RuntimeError(f"Missing target records: {missing}")

    wb.save(CURRENT_WORKBOOK)
    print(f"Workbook restored: {CURRENT_WORKBOOK}")
    print(f"Scores changed: {changed_scores}")
    print(f"Detail cells written: {detail_cells_written}")
    print(f"Score cells restored: {(ws.max_row - 1) * len(score_cols)}")


if __name__ == "__main__":
    restore()
