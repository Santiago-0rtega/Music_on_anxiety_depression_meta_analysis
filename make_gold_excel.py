import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import importlib.util, os, pandas as pd

spec = importlib.util.spec_from_file_location("gs", "build_gold_standard.py")
gs = importlib.util.module_from_spec(spec); spec.loader.exec_module(gs)
A, ITEMS, ITEM_NAME = gs.A, gs.ITEMS, gs.ITEM_NAME

# study order from roster
roster = pd.read_csv('data/crime_q_cohort_roster.csv')
studies = list(dict.fromkeys(roster['Study_ID'].tolist()))

SCORE_FILL = {
 'Yes':'C6EFCE','No':'FFC7CE','Partly':'FFEB9C','Unclear':'D9D9D9','NA':'F2F2F2'}
SCORE_FONT = {
 'Yes':'006100','No':'9C0006','Partly':'9C6500','Unclear':'3F3F3F','NA':'7F7F7F'}

thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin,right=thin,top=thin,bottom=thin)
hdr_fill = PatternFill('solid', fgColor='366092')
hdr_font = Font(bold=True, color='FFFFFF', size=10)
wrap_top = Alignment(wrap_text=True, vertical='top')

wb = openpyxl.Workbook(); wb.remove(wb.active)

# ---------- Sheet 1: Scores matrix (study x item) ----------
ws = wb.create_sheet('Scores')
ws.cell(1,1,'Study_ID').font=hdr_font; ws.cell(1,1).fill=hdr_fill
for c,it in enumerate(ITEMS, 2):
    cell=ws.cell(1,c,it); cell.font=hdr_font; cell.fill=hdr_fill
    cell.alignment=Alignment(wrap_text=True,vertical='center',horizontal='center')
for r,s in enumerate(studies, 2):
    ws.cell(r,1,s).alignment=wrap_top; ws.cell(r,1).border=border
    for c,it in enumerate(ITEMS,2):
        sc=A[s][it][0]
        cell=ws.cell(r,c,sc)
        cell.fill=PatternFill('solid',fgColor=SCORE_FILL.get(sc,'FFFFFF'))
        cell.font=Font(color=SCORE_FONT.get(sc,'000000'),size=10,bold=True)
        cell.alignment=Alignment(horizontal='center',vertical='center')
        cell.border=border
ws.column_dimensions['A'].width=34
for c in range(2,len(ITEMS)+2): ws.column_dimensions[get_column_letter(c)].width=7
ws.freeze_panes='B2'
ws.row_dimensions[1].height=30

# ---------- Sheet 2: Full assessment (long format) ----------
wl = wb.create_sheet('Full_Assessment')
heads=['Study_ID','Item','Item_Name','SCORE','JUSTIFICATION','VERBATIM']
for c,h in enumerate(heads,1):
    cell=wl.cell(1,c,h); cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=Alignment(vertical='center')
r=2
for s in studies:
    for it in ITEMS:
        sc,ju,ve=A[s][it]
        wl.cell(r,1,s); wl.cell(r,2,it); wl.cell(r,3,ITEM_NAME[it])
        scc=wl.cell(r,4,sc); scc.fill=PatternFill('solid',fgColor=SCORE_FILL.get(sc,'FFFFFF'))
        scc.font=Font(color=SCORE_FONT.get(sc,'000000'),bold=True)
        scc.alignment=Alignment(horizontal='center',vertical='top')
        wl.cell(r,5,ju); wl.cell(r,6,ve)
        for c in range(1,7): wl.cell(r,c).alignment=wrap_top; wl.cell(r,c).border=border
        r+=1
widths=[30,8,30,10,55,60]
for c,w in enumerate(widths,1): wl.column_dimensions[get_column_letter(c)].width=w
wl.freeze_panes='A2'

# ---------- Sheet 3: Wide template (study x 60 cols) mirroring NotebookLM output ----------
ww = wb.create_sheet('Wide_NotebookLM_Format')
heads=['Study_ID','Study_Title']
for it in ITEMS:
    nm=ITEM_NAME[it]
    heads += [f'{it}_{nm}_SCORE', f'{it}_{nm}_JUSTIFICATION', f'{it}_{nm}_VERBATIM']
for c,h in enumerate(heads,1):
    cell=ww.cell(1,c,h); cell.font=Font(bold=True,color='FFFFFF',size=9); cell.fill=hdr_fill
    cell.alignment=Alignment(wrap_text=True,vertical='center')
for r,s in enumerate(studies,2):
    ww.cell(r,1,s); ww.cell(r,2,s)
    c=3
    for it in ITEMS:
        sc,ju,ve=A[s][it]
        scc=ww.cell(r,c,sc); scc.fill=PatternFill('solid',fgColor=SCORE_FILL.get(sc,'FFFFFF'))
        scc.font=Font(color=SCORE_FONT.get(sc,'000000'),bold=True)
        ww.cell(r,c+1,ju); ww.cell(r,c+2,ve); c+=3
    for c2 in range(1,len(heads)+1): ww.cell(r,c2).alignment=wrap_top
ww.column_dimensions['A'].width=28; ww.column_dimensions['B'].width=28
for c in range(3,len(heads)+1):
    t=(c-3)%3; ww.column_dimensions[get_column_letter(c)].width=11 if t==0 else (34 if t==1 else 40)
ww.freeze_panes='C2'

# ---------- Sheet 4: Summary counts per item ----------
wsum=wb.create_sheet('Summary')
wsum.cell(1,1,'Item').font=hdr_font; wsum.cell(1,1).fill=hdr_fill
for c,lab in enumerate(['Item','Name','Yes','Partly','No','Unclear','NA'],1):
    cell=wsum.cell(1,c,lab); cell.font=hdr_font; cell.fill=hdr_fill
for r,it in enumerate(ITEMS,2):
    cnt={k:0 for k in ['Yes','Partly','No','Unclear','NA']}
    for s in studies: cnt[A[s][it][0]]=cnt.get(A[s][it][0],0)+1
    wsum.cell(r,1,it); wsum.cell(r,2,ITEM_NAME[it])
    for c,k in enumerate(['Yes','Partly','No','Unclear','NA'],3): wsum.cell(r,c,cnt[k])
wsum.column_dimensions['A'].width=8; wsum.column_dimensions['B'].width=32
for c in range(3,8): wsum.column_dimensions[get_column_letter(c)].width=9

out='data/MUSIC-CRIME-Q_GOLD-STANDARD_assessment.xlsx'
wb.save(out)
print('Saved', out)
print('Sheets:', wb.sheetnames)
print('Studies:', len(studies), '| Items:', len(ITEMS))
